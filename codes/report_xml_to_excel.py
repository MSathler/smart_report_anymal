import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# Load the XML file
tree = ET.parse("report.xml")
root = tree.getroot()
files = []
# Create an empty DataFrame
df = pd.DataFrame(columns=['Horário', 'Nome', 'Mensagem', 'Valor', 'Estado', 'Foto'])

# Process the entries in the XML
for entry in root.findall('entry'):
    level = entry.find('level').text
    if level == "INFO":
        time_of_day = entry.find('time_of_day').text
        message = entry.find('message').text
        if entry.find('file').text is not None:
            files.append(entry.find('file').text) # Save file paths
        name = message.split("'")[1]
        value, status = '-', '-'  # Default values for non-temperature cases

        if 'took image' in message:
            message_t = f'Imagem tirada com sucesso de: {name}'
        elif 'analyzed temperature' in message:
            status = entry.find('status').text
            value = entry.find('value').text
            message_t = f'Temperatura analisada com sucesso de {name}'
        else:
            continue  # Skip other messages

        # Create a temporary DataFrame for this iteration
        temp_df = pd.DataFrame({
            'Horário': [time_of_day],
            'Nome': [name],
            'Mensagem': [message_t],
            'Valor': [value],
            'Estado': [status],
            'Foto': [files[-1]]
        })

        # Append data to the main DataFrame
        df = pd.concat([df, temp_df], ignore_index=True)

# Save the DataFrame to an Excel file using pandas
df.to_excel('output2.xlsx', index=False, engine='openpyxl')

# Load the workbook and the sheet
wb = load_workbook('output2.xlsx')
ws = wb.active

# Set column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 62.5


# Apply styles to header
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws[f'{col}1']
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add images

for i in range(len(files)):  # Start from 2 to skip the header
    # Set column heights
    ws.row_dimensions[i+2].height = 240
    img = Image(files[i])
    img.width = 500  # Set the image width
    img.height = 300  # Set the image height
    ws.add_image(img, f'F{i+2}')  # Assuming images should be in column F

# Save the workbook
wb.save('report.xlsx')
